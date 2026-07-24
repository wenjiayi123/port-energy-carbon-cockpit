#!/usr/bin/env python3
"""Build the public hourly CarbonOps benchmark package.

The script joins five inspectable public-data layers:

1. Port of Los Angeles official monthly container statistics;
2. U.S. EIA California commercial monthly retail-price anchors;
3. U.S. EIA-930 LADWP hourly load and consumed-carbon intensity;
4. U.S. EPA eGRID CAMX output CO2e rate for the absolute carbon level;
5. LADWP published commercial time-of-use periods, or optional CAISO OASIS
   SP15 day-ahead LMPs, for the within-month price shape.

Monthly throughput remains an observed anchor. Hourly terminal demand is a
declared deterministic allocation, not observed TOS telemetry. Hourly retail
price and carbon series are likewise transparent scenario inputs: the retail
price level is anchored to EIA monthly means and its intraday shape is a
declared multiplier over LADWP's published time bands; hourly carbon uses EIA's
published consumed-electricity intensity.
"""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
from datetime import datetime, timezone
import hashlib
import io
import json
from pathlib import Path
import ssl
import time
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen
import zipfile

import certifi
import numpy as np
from openpyxl import load_workbook
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
DATASET_DIR = ROOT / "backend" / "app" / "data" / "datasets"
MONTHLY_DATASET = DATASET_DIR / "port_la_2020_2025_monthly.csv"
MONTHLY_METADATA = MONTHLY_DATASET.with_suffix(".metadata.json")
DEFAULT_OUTPUT = DATASET_DIR / "port_la_2020_2025_hourly.csv"
CAISO_ENDPOINT = "https://oasis.caiso.com/oasisapi/SingleZip"
EIA_LDWP_WORKBOOK_URL = (
    "https://www.eia.gov/electricity/gridmonitor/knownissues/xls/LDWP.xlsx"
)
LADWP_COMMERCIAL_RATES_URL = (
    "https://www.ladwp.com/account/understanding-your-rates/commercial-electric-rates"
)
CAISO_NODE = "TH_SP15_GEN-APND"
FX_CNY_PER_USD = 7.20
HOURLY_DEMAND_PROFILE = np.array(
    [
        0.72,
        0.66,
        0.62,
        0.60,
        0.64,
        0.78,
        0.92,
        1.08,
        1.18,
        1.22,
        1.17,
        1.10,
        1.04,
        1.08,
        1.16,
        1.24,
        1.30,
        1.25,
        1.14,
        1.02,
        0.94,
        0.88,
        0.82,
        0.76,
    ],
    dtype=np.float64,
)
HOURLY_DEMAND_PROFILE /= HOURLY_DEMAND_PROFILE.mean()
LB_TO_KG = 0.45359237


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def fetch_bytes(url: str, *, attempts: int = 5) -> bytes:
    last_error: Exception | None = None
    ssl_context = ssl.create_default_context(cafile=certifi.where())
    for attempt in range(attempts):
        try:
            request = Request(url, headers={"User-Agent": "CarbonOps-public-benchmark/3.0"})
            with urlopen(request, timeout=90, context=ssl_context) as response:
                return response.read()
        except Exception as exc:  # pragma: no cover - exercised only on network failure
            last_error = exc
            time.sleep(min(20.0, 2.0 ** attempt))
    raise RuntimeError(f"Public-data request failed after {attempts} attempts: {url}") from last_error


def month_windows(start_year: int, end_year: int) -> list[tuple[datetime, datetime]]:
    windows: list[tuple[datetime, datetime]] = []
    for year in range(start_year, end_year + 1):
        for month in range(1, 13):
            start = datetime(year, month, 1, 8, tzinfo=timezone.utc)
            end = (
                datetime(year + 1, 1, 1, 8, tzinfo=timezone.utc)
                if month == 12
                else datetime(year, month + 1, 1, 8, tzinfo=timezone.utc)
            )
            windows.append((start, end))
    return windows


def caiso_url(start: datetime, end: datetime) -> str:
    parameters = {
        "resultformat": 6,
        "queryname": "PRC_LMP",
        "version": 12,
        "startdatetime": start.strftime("%Y%m%dT%H:%M-0000"),
        "enddatetime": end.strftime("%Y%m%dT%H:%M-0000"),
        "market_run_id": "DAM",
        "node": CAISO_NODE,
    }
    return f"{CAISO_ENDPOINT}?{urlencode(parameters)}"


def fetch_caiso_window(start: datetime, end: datetime) -> tuple[pd.DataFrame, dict[str, Any]]:
    url = caiso_url(start, end)
    payload = fetch_bytes(url)
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            names = [name for name in archive.namelist() if name.lower().endswith(".csv")]
            if len(names) != 1:
                raise ValueError(f"Expected one CAISO CSV, found {names}")
            csv_bytes = archive.read(names[0])
    except zipfile.BadZipFile as exc:
        raise RuntimeError(f"CAISO returned a non-ZIP payload for {start:%Y-%m}") from exc
    frame = pd.read_csv(io.BytesIO(csv_bytes))
    frame = frame[
        (frame["LMP_TYPE"].astype(str) == "LMP")
        & (frame["NODE"].astype(str) == CAISO_NODE)
    ][["INTERVALSTARTTIME_GMT", "MW"]].copy()
    frame["timestamp"] = pd.to_datetime(frame["INTERVALSTARTTIME_GMT"], utc=True)
    frame["wholesale_lmp_usd_per_mwh"] = pd.to_numeric(frame["MW"], errors="raise")
    frame = frame[["timestamp", "wholesale_lmp_usd_per_mwh"]]
    return frame, {
        "period": start.strftime("%Y-%m"),
        "url": url,
        "response_sha256": sha256_bytes(payload),
        "rows": int(len(frame)),
    }


def fetch_caiso(start_year: int, end_year: int, workers: int) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    frames: list[pd.DataFrame] = []
    evidence: list[dict[str, Any]] = []
    windows = month_windows(start_year, end_year)
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        futures = {
            executor.submit(fetch_caiso_window, start, end): (start, end)
            for start, end in windows
        }
        for future in as_completed(futures):
            frame, item = future.result()
            frames.append(frame)
            evidence.append(item)
            print(f"CAISO {item['period']}: {item['rows']} LMP rows", flush=True)
    result = pd.concat(frames, ignore_index=True).drop_duplicates("timestamp", keep="last")
    return result.sort_values("timestamp").reset_index(drop=True), sorted(
        evidence, key=lambda item: item["period"]
    )


def fetch_eia_hourly_workbook(
    start_year: int,
    end_year: int,
    workbook_cache: Path | None = None,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    payload = (
        workbook_cache.read_bytes()
        if workbook_cache is not None and workbook_cache.exists()
        else fetch_bytes(EIA_LDWP_WORKBOOK_URL)
    )
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    worksheet = workbook["Published Hourly Data"]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    index = {str(value): position for position, value in enumerate(header)}
    required = {
        "UTC time",
        "Adjusted demand",
        "Demand",
        "Consumed Electricity",
        "CO2 Emissions Intensity for Consumed Electricity",
    }
    missing = required - set(index)
    if missing:
        raise RuntimeError(f"EIA-930 workbook is missing columns: {sorted(missing)}")
    records: list[dict[str, Any]] = []
    for row in rows:
        timestamp = row[index["UTC time"]]
        if not isinstance(timestamp, datetime) or not start_year <= timestamp.year <= end_year:
            continue
        demand = row[index["Adjusted demand"]]
        if demand is None:
            demand = row[index["Demand"]]
        consumed = row[index["Consumed Electricity"]]
        intensity_lb = row[index["CO2 Emissions Intensity for Consumed Electricity"]]
        if demand is None or intensity_lb is None:
            continue
        records.append(
            {
                "timestamp": timestamp.replace(tzinfo=timezone.utc),
                "eia930_demand_mw": max(0.0, float(demand)),
                "eia930_consumed_electricity_mwh": max(
                    0.0, float(consumed if consumed is not None else demand)
                ),
                "eia930_consumed_carbon_lb_per_kwh": max(0.0, float(intensity_lb)),
                "grid_carbon_kg_per_kwh": max(0.0, float(intensity_lb)) * LB_TO_KG,
            }
        )
    frame = pd.DataFrame(records).drop_duplicates("timestamp", keep="last")
    frame = frame.sort_values("timestamp").reset_index(drop=True)
    expected_hours = int(
        (
            datetime(end_year + 1, 1, 1, tzinfo=timezone.utc)
            - datetime(start_year, 1, 1, tzinfo=timezone.utc)
        ).total_seconds()
        // 3600
    )
    coverage = len(frame) / max(1, expected_hours)
    if coverage < 0.98:
        raise RuntimeError(
            f"EIA-930 hourly coverage is {coverage:.2%}; at least 98% is required"
        )
    observed_rows = len(frame)
    full_index = pd.date_range(
        start=f"{start_year}-01-01T00:00:00Z",
        end=f"{end_year + 1}-01-01T00:00:00Z",
        inclusive="left",
        freq="h",
    )
    frame = frame.set_index("timestamp").reindex(full_index)
    frame.index.name = "timestamp"
    observed_mask = frame["grid_carbon_kg_per_kwh"].notna() & frame[
        "eia930_demand_mw"
    ].notna()
    calendar_key = pd.MultiIndex.from_arrays(
        [frame.index.month, frame.index.hour], names=["month", "hour"]
    )
    for column in (
        "eia930_demand_mw",
        "eia930_consumed_electricity_mwh",
        "eia930_consumed_carbon_lb_per_kwh",
        "grid_carbon_kg_per_kwh",
    ):
        observed = frame.loc[observed_mask, column]
        observed_key = pd.MultiIndex.from_arrays(
            [observed.index.month, observed.index.hour], names=["month", "hour"]
        )
        medians = observed.groupby(observed_key).median()
        fallback = pd.Series(calendar_key.map(medians), index=frame.index, dtype=float)
        frame[column] = frame[column].fillna(fallback).fillna(float(observed.median()))
    frame["eia930_quality_code"] = np.where(
        observed_mask, "reported", "calendar-hour-median-imputed"
    )
    frame = frame.reset_index()
    print(
        f"EIA-930 LDWP workbook: {observed_rows}/{expected_hours} reported hours "
        f"({coverage:.2%}); {expected_hours - observed_rows} explicitly imputed",
        flush=True,
    )
    return frame, [
        {
            "url": EIA_LDWP_WORKBOOK_URL,
            "cache_path": str(workbook_cache) if workbook_cache else None,
            "response_sha256": sha256_bytes(payload),
            "source_rows": int(worksheet.max_row - 1),
            "reported_rows": int(observed_rows),
            "imputed_rows": int(expected_hours - observed_rows),
            "selected_rows": int(len(frame)),
            "expected_hours": expected_hours,
            "coverage": round(coverage, 6),
        }
    ]


def normalized_shape(values: pd.Series, *, low: float = 0.55, high: float = 1.65) -> pd.Series:
    p05 = float(values.quantile(0.05))
    p95 = float(values.quantile(0.95))
    if p95 <= p05:
        return pd.Series(np.ones(len(values)), index=values.index)
    scaled = ((values - p05) / (p95 - p05)).clip(lower=0.0, upper=1.0)
    multiplier = low + (high - low) * scaled
    return multiplier / max(float(multiplier.mean()), 1e-9)


def build_hourly_dataset(
    monthly: pd.DataFrame,
    caiso: pd.DataFrame | None,
    fuel_mix: pd.DataFrame,
) -> pd.DataFrame:
    joined = fuel_mix.copy()
    if caiso is not None:
        joined = caiso.merge(joined, on="timestamp", how="inner", validate="one_to_one")
    joined["period"] = joined["timestamp"].dt.strftime("%Y-%m")
    joined = joined.merge(
        monthly[
            [
                "period",
                "split",
                "loaded_import_teu",
                "loaded_export_teu",
                "total_teu",
                "electricity_price_per_kwh",
                "fuel_price_per_liter",
            ]
        ].rename(
            columns={
                "total_teu": "monthly_total_teu",
                "loaded_import_teu": "monthly_loaded_import_teu",
                "loaded_export_teu": "monthly_loaded_export_teu",
                "electricity_price_per_kwh": "monthly_eia_price_cny_per_kwh",
            }
        ),
        on="period",
        how="inner",
        validate="many_to_one",
    )
    if joined.empty:
        raise RuntimeError("No overlapping hourly records were found")
    output_frames: list[pd.DataFrame] = []
    for period, month in joined.groupby("period", sort=True):
        month = month.sort_values("timestamp").copy()
        if "wholesale_lmp_usd_per_mwh" in month:
            month["price_shape"] = normalized_shape(
                month["wholesale_lmp_usd_per_mwh"]
            )
        else:
            local_time = month["timestamp"].dt.tz_convert("America/Los_Angeles")
            weekday = local_time.dt.dayofweek < 5
            hour = local_time.dt.hour
            # LADWP publishes base, low-peak and high-peak clock bands, but the
            # page does not expose one universal terminal tariff. These
            # predeclared relative multipliers preserve those official bands
            # and are rescaled to the EIA monthly commercial mean below.
            tou_shape = np.full(len(month), 0.75, dtype=np.float64)
            tou_shape[weekday & hour.between(10, 12)] = 1.15
            tou_shape[weekday & hour.between(13, 16)] = 1.45
            tou_shape[weekday & hour.between(17, 19)] = 1.15
            month["price_shape"] = tou_shape / float(tou_shape.mean())
        month["electricity_price_per_kwh"] = (
            month["monthly_eia_price_cny_per_kwh"] * month["price_shape"]
        )
        month["grid_carbon_kg_per_kwh"] = month["grid_carbon_kg_per_kwh"].clip(
            lower=0.03, upper=1.20
        )
        hour_local = month["timestamp"].dt.tz_convert("America/Los_Angeles").dt.hour
        profile = pd.Series(HOURLY_DEMAND_PROFILE[hour_local.to_numpy()], index=month.index)
        profile /= max(float(profile.sum()), 1e-9)
        month["total_teu"] = month["monthly_total_teu"] * profile
        month["loaded_import_teu"] = (
            month["total_teu"]
            * month["monthly_loaded_import_teu"]
            / month["monthly_total_teu"].clip(lower=1e-9)
        )
        month["loaded_export_teu"] = (
            month["total_teu"]
            * month["monthly_loaded_export_teu"]
            / month["monthly_total_teu"].clip(lower=1e-9)
        )
        output_frames.append(month)
        print(f"joined {period}: {len(month)} hourly rows", flush=True)
    result = pd.concat(output_frames, ignore_index=True)
    result["observation_hours"] = 1.0
    result["source_id"] = result["timestamp"].dt.strftime(
        "PORT-LA+EIA+EGRID+EIA930:%Y-%m-%dT%H"
    )
    result["timestamp_utc"] = result["timestamp"].dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    columns = [
        "timestamp_utc",
        "period",
        "split",
        "loaded_import_teu",
        "loaded_export_teu",
        "total_teu",
        "grid_carbon_kg_per_kwh",
        "electricity_price_per_kwh",
        "fuel_price_per_liter",
        "eia930_demand_mw",
        "eia930_consumed_electricity_mwh",
        "eia930_consumed_carbon_lb_per_kwh",
        "eia930_quality_code",
        "monthly_eia_price_cny_per_kwh",
        "monthly_total_teu",
        "observation_hours",
        "source_id",
    ]
    if "wholesale_lmp_usd_per_mwh" in result:
        columns.insert(9, "wholesale_lmp_usd_per_mwh")
    return result[columns].sort_values("timestamp_utc").reset_index(drop=True)


def write_package(
    frame: pd.DataFrame,
    output: Path,
    monthly_metadata: dict[str, Any],
    caiso_evidence: list[dict[str, Any]],
    eia_evidence: list[dict[str, Any]],
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(output, index=False, float_format="%.8f", quoting=csv.QUOTE_MINIMAL)
    output_sha = sha256_bytes(output.read_bytes())
    metadata = {
        "id": output.stem,
        "name": (
            "Port of Los Angeles 2020-2025 hourly dispatch benchmark: official monthly "
            "throughput and retail-price anchors, EIA-930 LADWP hourly generation mix, "
            "EPA eGRID CAMX calibration, and optional CAISO SP15 DAM LMP"
        ),
        "version": "2020-2025.3",
        "license": monthly_metadata["license"],
        "source_urls": [
            *monthly_metadata["source_urls"],
            "https://api.eia.gov/v2/electricity/rto/fuel-type-data/",
            "https://www.eia.gov/electricity/gridmonitor/about",
            LADWP_COMMERCIAL_RATES_URL,
            *(
                [
                    "https://oasis.caiso.com/",
                    "https://www.caiso.com/documents/oasis-frequently-asked-questions.pdf",
                ]
                if caiso_evidence
                else []
            ),
        ],
        "attribution": (
            "Port of Los Angeles; U.S. EIA; U.S. EPA"
            + ("; California Independent System Operator" if caiso_evidence else "")
            + ". See fetch evidence and transformation notes."
        ),
        "scope_note": (
            "Offline public-data benchmark. Throughput is observed monthly and allocated "
            "to hours by a disclosed profile; hourly retail price and carbon are calibrated "
            "proxies, not terminal invoices, marginal-emission signals, or TOS/EMS telemetry."
        ),
        "temporal_mode": "sequential_rows",
        "time_column": "timestamp_utc",
        "evaluation_episode_limit": 48,
        "evaluation_sampling": (
            "48 deterministic, uniformly spaced 24-hour episodes spanning each "
            "validation/test year; start indices are persisted in benchmark evidence"
        ),
        "units": {
            "loaded_import_teu": "TEU/hour modeled allocation",
            "loaded_export_teu": "TEU/hour modeled allocation",
            "total_teu": "TEU/hour modeled allocation",
            "grid_carbon_kg_per_kwh": "kgCO2e/kWh calibrated hourly proxy",
            "electricity_price_per_kwh": "CNY/kWh calibrated hourly proxy",
            "fuel_price_per_liter": "CNY/liter scenario",
            "wholesale_lmp_usd_per_mwh": "USD/MWh official CAISO DAM LMP",
            "eia930_demand_mw": "MW official EIA-930 adjusted/reported demand",
            "eia930_consumed_electricity_mwh": "MWh official EIA-930 consumed electricity",
            "eia930_consumed_carbon_lb_per_kwh": "lbCO2/kWh official EIA estimated consumed intensity",
            "eia930_quality_code": "reported or declared calendar-hour median imputation",
            "monthly_eia_price_cny_per_kwh": "CNY/kWh EIA monthly anchor",
            "monthly_total_teu": "TEU/month official Port total",
            "observation_hours": "hour",
        },
        "assumptions": [
            "Port monthly TEU is allocated across hours by a deterministic normalized profile.",
            (
                "CAISO SP15 day-ahead LMP supplies within-month price rank and volatility."
                if caiso_evidence
                else (
                    "LADWP published commercial time-of-use bands supply the "
                    "intraday price periods; relative multipliers are disclosed assumptions."
                )
            ),
            "Each hourly price shape is rescaled to the official EIA California commercial monthly mean.",
            "EIA-930 LADWP published hourly consumed-electricity CO2 intensity supplies the carbon signal.",
            "EPA eGRID CAMX remains an annual cross-check; the hourly signal is not rescaled to it.",
            "Missing EIA-930 hours are explicitly quality-coded and filled with the observed month-hour median.",
            "Storage and terminal physical parameters remain declared benchmark assumptions in metadata.",
        ],
        "intended_use": (
            "Chronological train/validation/test benchmarking of constrained energy-carbon "
            "dispatch, storage, shore power, and equipment-resource policies."
        ),
        "environment_parameters": {
            **monthly_metadata["environment_parameters"],
            # Capacities represent the full available fleet. Control ratios are
            # active-fleet fractions in [0.60, 1.00], never over-speed factors.
            "crane_capacity_teu_per_hour": 1_850.0,
            "yard_capacity_teu_per_hour": 2_050.0,
            "battery_capacity_kwh": 18_000.0,
            "battery_power_kw": 5_000.0,
            "battery_initial_soc": 0.50,
            "battery_min_soc": 0.10,
            "battery_max_soc": 0.90,
            "battery_charge_efficiency": 0.95,
            "battery_discharge_efficiency": 0.95,
            "battery_degradation_cny_per_kwh": 0.18,
            "terminal_soc_tolerance": 0.05,
        },
        "transformation": {
            "price": (
                "monthly_eia_price * normalized clipped "
                + (
                    "CAISO within-month LMP"
                    if caiso_evidence
                    else "LADWP commercial time-of-use band"
                )
                + (
                    " shape; CAISO values are clipped to 0.55-1.65"
                    if caiso_evidence
                    else " shape using disclosed base/low/high multipliers "
                    "0.75/1.15/1.45 before monthly mean rescaling"
                )
            ),
            "carbon": (
                "official EIA-930 hourly consumed-electricity intensity converted "
                "from lbCO2/kWh to kgCO2/kWh; EPA eGRID CAMX retained as annual cross-check"
            ),
            "throughput": (
                "official monthly total * normalized disclosed local-hour demand profile"
            ),
        },
        "public_source_evidence": {
            "monthly_package_sha256": sha256_bytes(
                MONTHLY_DATASET.read_bytes() + b"\0" + MONTHLY_METADATA.read_bytes()
            ),
            "caiso_response_count": len(caiso_evidence),
            "caiso_responses": caiso_evidence,
            "eia930_response_count": len(eia_evidence),
            "eia930_responses": eia_evidence,
        },
        "quality": {
            "rows": int(len(frame)),
            "start": str(frame["timestamp_utc"].min()),
            "end": str(frame["timestamp_utc"].max()),
            "missing_cells": int(frame.isna().sum().sum()),
            "duplicate_timestamps": int(frame["timestamp_utc"].duplicated().sum()),
            "csv_sha256": output_sha,
        },
    }
    metadata_path = output.with_suffix(".metadata.json")
    metadata_path.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "dataset": str(output),
                "metadata": str(metadata_path),
                "rows": len(frame),
                "csv_sha256": output_sha,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--start-year", type=int, default=2020)
    parser.add_argument("--end-year", type=int, default=2025)
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--include-caiso",
        action="store_true",
        help="Optionally fetch CAISO SP15 DAM LMP instead of using EIA-930 load shape.",
    )
    parser.add_argument(
        "--caiso-cache",
        type=Path,
        help="Optional parquet cache for an already downloaded CAISO frame.",
    )
    parser.add_argument(
        "--eia-cache",
        type=Path,
        help="Optional parquet cache for an already downloaded EIA-930 frame.",
    )
    parser.add_argument(
        "--eia-workbook-cache",
        type=Path,
        help="Optional local copy of the official LDWP.xlsx download.",
    )
    args = parser.parse_args()
    monthly = pd.read_csv(MONTHLY_DATASET)
    monthly_metadata = json.loads(MONTHLY_METADATA.read_text(encoding="utf-8"))
    caiso: pd.DataFrame | None = None
    caiso_evidence: list[dict[str, Any]] = []
    if args.include_caiso and args.caiso_cache and args.caiso_cache.exists():
        caiso = pd.read_parquet(args.caiso_cache)
        caiso_evidence = [{"cache": str(args.caiso_cache), "rows": len(caiso)}]
    elif args.include_caiso:
        caiso, caiso_evidence = fetch_caiso(
            args.start_year, args.end_year, args.workers
        )
        if args.caiso_cache:
            args.caiso_cache.parent.mkdir(parents=True, exist_ok=True)
            caiso.to_parquet(args.caiso_cache, index=False)
    if args.eia_cache and args.eia_cache.exists():
        fuel_mix = pd.read_parquet(args.eia_cache)
        eia_evidence = [{"cache": str(args.eia_cache), "rows": len(fuel_mix)}]
    else:
        fuel_mix, eia_evidence = fetch_eia_hourly_workbook(
            args.start_year, args.end_year, args.eia_workbook_cache
        )
        if args.eia_cache:
            args.eia_cache.parent.mkdir(parents=True, exist_ok=True)
            fuel_mix.to_parquet(args.eia_cache, index=False)
    frame = build_hourly_dataset(monthly, caiso, fuel_mix)
    write_package(frame, args.output, monthly_metadata, caiso_evidence, eia_evidence)


if __name__ == "__main__":
    main()
